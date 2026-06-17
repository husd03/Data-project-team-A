"""
Report generator.

Z výsledků agenta vytvoří:
  1. Excel soubor pro bankovní systém (strukturovaný, přehledný)
  2. CSV soubor pro import do CRM / datového skladu
  3. JSON summary pro logy a monitoring
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COL_HEADER_BG = "1F4E79"
COL_HEADER_FG = "FFFFFF"
COL_HIGH_BG   = "E2EFDA"
COL_MED_BG    = "FFF2CC"
COL_LOW_BG    = "F2F2F2"
COL_CASH_BG   = "EAF3DE"
COL_SAV_BG    = "DEEAF1"
COL_INV_BG    = "EDE7F6"


def generate_reports(
    results: pd.DataFrame,
    output_dir: str | Path,
    run_date: str,
) -> dict[str, Path]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    tag = run_date.replace("-", "")
    paths = {}
    paths["excel"] = _write_excel(results, out / f"odmeny_{tag}.xlsx", run_date)
    paths["csv"]   = _write_csv(results,   out / f"odmeny_{tag}.csv")
    paths["json"]  = _write_summary(results, out / f"summary_{tag}.json", run_date)
    return paths


# ── Definice sloupců ───────────────────────────────────────────────────────

def _col_headers() -> list[tuple[str, str, int]]:
    """(nadpis v Excelu, klíč v DataFrame, šířka sloupce)"""
    return [
        ("ID zákazníka",              "customer_id",       12),
        ("Segment",                   "segment",            12),
        ("Priorita",                  "priority_tier",      10),
        ("Skóre konverze",            "conversion_score",   14),
        ("Byl MAIN",                  "was_main_before",    10),
        ("Výzva 1 (kód)",             "vyzva_1",             9),
        ("Výzva 1",                   "vyzva_1_nazev",      34),
        ("Výzva 1 - důvod",           "vyzva_1_duvod",      34),
        ("Výzva 2 (kód)",             "vyzva_2",             9),
        ("Výzva 2",                   "vyzva_2_nazev",      34),
        ("Výzva 2 - důvod",           "vyzva_2_duvod",      34),
        ("Výzva 3 (kód)",             "vyzva_3",             9),
        ("Výzva 3",                   "vyzva_3_nazev",      34),
        ("Výzva 3 - důvod",           "vyzva_3_duvod",      34),
        ("Příjem splněn (podíl)",     "mesice_prijem_z",    16),
        ("Transakce splněny (podíl)", "mesice_trx_z",       18),
        ("Avg příjem (Kč)",           "avg_cr_czk",         16),
        ("Avg transakce",             "avg_trx",            14),
        ("Avg SPB login",             "avg_spb",            14),
        ("Avg DC platby",             "avg_dc",             13),
        ("Avg zůstatek (Kč)",         "avg_balance_czk",    18),
        ("Věk",                       "age",                 8),
        ("Pohlaví",                   "gender",              9),
        ("Zákazník od roku",          "year_joined",        14),
        ("Datum spuštění",            "run_date",           14),
    ]


# ── Excel ──────────────────────────────────────────────────────────────────

def _write_excel(df: pd.DataFrame, path: Path, run_date: str) -> Path:
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Vsichni SECONDARY"
    _fill_sheet(ws_all, df, run_date, "Doporučení odměn pro SECONDARY zákazníky")

    high = df[df["priority_tier"] == "HIGH"].copy()
    ws_h = wb.create_sheet("HIGH priorita")
    _fill_sheet(ws_h, high, run_date, "Zákazníci s nejvyšší šancí konverze — HIGH")

    ws_s = wb.create_sheet("Souhrn")
    _fill_summary(ws_s, df, run_date)

    wb.save(path)
    return path


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, bottom=s)


def _fill_sheet(ws, df: pd.DataFrame, run_date: str, title: str) -> None:
    headers = _col_headers()
    ncols   = len(headers)
    brd     = _border()

    # Řádek 1: titulek
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = f"{title} — {run_date}"
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=COL_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Řádek 2: záhlaví sloupců
    for ci, (label, _, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 20

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

    # Barvy řádků
    tier_bg   = {"HIGH": COL_HIGH_BG, "MEDIUM": COL_MED_BG, "LOW": COL_LOW_BG}
    challenge_bg = {"C1": "FFF2CC", "C2": "EAF3DE", "C3": "DEEAF1", "C4": "EDE7F6", "C5": "FCE4D6", "C6": "E2EFDA"}
    col_keys  = [k for _, k, _ in headers]

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        tier   = str(row.get("priority_tier", ""))
        row_bg = tier_bg.get(tier, "FFFFFF")

        for ci, key in enumerate(col_keys, 1):
            val  = _clean(row.get(key, ""))
            cell = ws.cell(row=ri, column=ci, value=val)
            c1_code = str(row.get("vyzva_1",""))
            bg = challenge_bg.get(c1_code, row_bg) if key == "vyzva_1_nazev" else row_bg
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.font   = Font(name="Arial", size=10)
            cell.border = brd


def _fill_summary(ws, df: pd.DataFrame, run_date: str) -> None:
    # Titulek
    ws["A1"] = f"Souhrn — {run_date}"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16

    # Hodnoty — vše převedeno na čisté Python int/float/str
    rows = [
        ("Celkem SECONDARY zákazníků",  int(len(df))),
        (None, None),
        ("PRIORITA", None),
        ("HIGH",   int((df["priority_tier"] == "HIGH").sum())),
        ("MEDIUM", int((df["priority_tier"] == "MEDIUM").sum())),
        ("LOW",    int((df["priority_tier"] == "LOW").sum())),
        (None, None),
        ("VÝZVY (počet zákazníků)", None),
        ("C1 - příjem 15 000 Kč", int(((df["vyzva_1"]=="C1")|(df["vyzva_2"]=="C1")|(df["vyzva_3"]=="C1")).sum())),
        ("C2 - 10x kartou",        int(((df["vyzva_1"]=="C2")|(df["vyzva_2"]=="C2")|(df["vyzva_3"]=="C2")).sum())),
        ("C3 - energie/telco",     int(((df["vyzva_1"]=="C3")|(df["vyzva_2"]=="C3")|(df["vyzva_3"]=="C3")).sum())),
        ("C4 - předplatné",        int(((df["vyzva_1"]=="C4")|(df["vyzva_2"]=="C4")|(df["vyzva_3"]=="C4")).sum())),
        ("C5 - SPB aktivita",      int(((df["vyzva_1"]=="C5")|(df["vyzva_2"]=="C5")|(df["vyzva_3"]=="C5")).sum())),
        ("C6 - investice",         int(((df["vyzva_1"]=="C6")|(df["vyzva_2"]=="C6")|(df["vyzva_3"]=="C6")).sum())),
        (None, None),
        ("ZÁKAZNÍCI CO BYLI MAIN", None),
        ("Byl MAIN dříve",    int(df["was_main_before"].sum())),
        ("Nikdy nebyl MAIN",  int((~df["was_main_before"]).sum())),
        (None, None),
        ("PRŮMĚRNÉ HODNOTY", None),
        ("Zákazníků s příjmem 0 (žádný splněný měs.)", int((df["mesice_prijem_z"] == 0.0).sum())),
        ("Průměrný conversion score",   float(round(df["conversion_score"].mean(), 1))),
    ]

    for ri, (label, val) in enumerate(rows, 3):
        if label is None:
            continue

        label_cell = ws.cell(row=ri, column=1, value=label)
        is_section = (val is None)
        label_cell.font = Font(name="Arial", size=10, bold=is_section)

        if val is not None:
            val_cell = ws.cell(row=ri, column=2, value=val)
            val_cell.font = Font(name="Arial", size=10)


# ── CSV ────────────────────────────────────────────────────────────────────

def _write_csv(df: pd.DataFrame, path: Path) -> Path:
    col_map = {key: label for label, key, _ in _col_headers()}
    keys    = [key for _, key, _ in _col_headers()]
    out_df  = df[keys].rename(columns=col_map)
    out_df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


# ── JSON summary ───────────────────────────────────────────────────────────

def _write_summary(df: pd.DataFrame, path: Path, run_date: str) -> Path:
    summary = {
        "run_date":             run_date,
        "generated_at":         datetime.now().isoformat(timespec="seconds"),
        "total_secondary":      int(len(df)),
        "priority_counts": {
            "HIGH":   int((df["priority_tier"] == "HIGH").sum()),
            "MEDIUM": int((df["priority_tier"] == "MEDIUM").sum()),
            "LOW":    int((df["priority_tier"] == "LOW").sum()),
        },
        "challenge_counts": {
            "C1": int(((df["vyzva_1"]=="C1")|(df["vyzva_2"]=="C1")|(df["vyzva_3"]=="C1")).sum()),
            "C2": int(((df["vyzva_1"]=="C2")|(df["vyzva_2"]=="C2")|(df["vyzva_3"]=="C2")).sum()),
            "C3": int(((df["vyzva_1"]=="C3")|(df["vyzva_2"]=="C3")|(df["vyzva_3"]=="C3")).sum()),
            "C4": int(((df["vyzva_1"]=="C4")|(df["vyzva_2"]=="C4")|(df["vyzva_3"]=="C4")).sum()),
            "C5": int(((df["vyzva_1"]=="C5")|(df["vyzva_2"]=="C5")|(df["vyzva_3"]=="C5")).sum()),
            "C6": int(((df["vyzva_1"]=="C6")|(df["vyzva_2"]=="C6")|(df["vyzva_3"]=="C6")).sum()),
        },
        "was_main_before":      int(df["was_main_before"].sum()),
        "avg_conversion_score": float(round(df["conversion_score"].mean(), 2)),
        "zakazniku_prijem_0": int((df["mesice_prijem_z"] == 0.0).sum()),
    }
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


# ── Helper ─────────────────────────────────────────────────────────────────

def _clean(val):
    """Převede hodnoty na typy které openpyxl bezpečně zapíše."""
    if isinstance(val, bool):
        return "Ano" if val else "Ne"
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return val
